
import os
import pdfkit
import openpyxl
import platform
from datetime import datetime
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

from Cheese.httpClientErrors import *
from Cheese.cheeseController import CheeseController as cc
from Cheese.resourceManager import ResMan
from Cheese.appSettings import Settings

from src.repositories.eventsRepository import EventsRepository
from src.repositories.clubsRepository import ClubsRepository

from src.other.billCalculator import BillCalculator

#@controller /bills;
class BillsController(cc):

    medium_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

    medium_left = Border(left=Side(style='medium'))
    medium_right = Border(right=Side(style='medium'))
    medium_right_thin_left = Border(right=Side(style='medium'), left=Side(style='medium'))

    #@post /getBillXlsx;
    @staticmethod
    def getBillXlsx(server, path, auth):
        args = cc.readArgs(server)

        cc.checkJson(["JBS", "ARRIVALS", "DEPARTS", "EVENT_ID"], args)

        if (len(args["JBS"]) == 0):
            raise BadRequest("There not any people")

        event = EventsRepository.find(args["EVENT_ID"])
        club = ClubsRepository.find(args["JBS"][0]["CLUB_ID"])

        bad, bpd, bsd = BillCalculator.getCalculatedBillData(args)

        billName = BillsController.createXlsx(bad, bpd, bsd, event, club)

        return cc.createResponse({"BILL": billName})

    #@post /getBillPdf;
    @staticmethod
    def getBillPdf(server, path, auth):
        args = cc.readArgs(server)

        cc.checkJson(["JBS", "ARRIVALS", "DEPARTS", "EVENT_ID"], args)

        if (len(args["JBS"]) == 0):
            raise BadRequest("There not any people")

        event = EventsRepository.find(args["EVENT_ID"])
        club = ClubsRepository.find(args["JBS"][0]["CLUB_ID"])

        bad, bpd, bsd = BillCalculator.getCalculatedBillData(args)

        billName = BillsController.createPdf(bad, bpd, bsd, event, club)

        return cc.createResponse({"BILL": billName})

    #METHOD

    @staticmethod
    def createXlsx(bad, bpd, bsd, event, club):
        wb = openpyxl.load_workbook(ResMan.resources("xlsxTemp.xlsx"))
        sheet = wb.active

        sheet.cell(11, 2, event.name)
        sheet.cell(11, 7, event.place.upper() + " " + event.event_start.strftime("%Y"))
        sheet.cell(14, 4, "NEZAPOMENOUT")
        sheet.cell(14, 7, datetime.now().strftime("%d.%m.%Y"))
        sheet.cell(15, 4, club.name)

        sheet = BillsController.emptyRow(sheet, 20)
        row = 0
        for i, room in enumerate(bad["ROOMS"]):
            row = i+21
            sheet.cell(row, 2, room["room_name"]).border = BillsController.medium_left
            sheet.cell(row, 3, room["start_date"]).alignment = Alignment(horizontal="center")
            sheet.cell(row, 4, room["end_date"]).alignment = Alignment(horizontal="center")
            sheet.cell(row, 5, room["count_room"]).alignment = Alignment(horizontal="center")
            sheet.cell(row, 6, room["count_people"]).alignment = Alignment(horizontal="center")
            sheet.cell(row, 7, room["nights"]).alignment = Alignment(horizontal="center")
            sheet.cell(row, 8, f"{room['price_ro']} €").alignment = Alignment(horizontal="center")
            sheet.cell(row, 9, f"{room['total']} €").border = BillsController.medium_right_thin_left
            sheet.cell(row, 9, f"{room['total']} €").alignment = Alignment(horizontal="right")
            sheet.cell(row, 9, f"{room['total']} €").font = Font(bold=True)

        row += 1
        sheet = BillsController.emptyRow(sheet, row)

        for item in bsd["ITEMS"]:
            row += 1
            sheet = BillsController.createTotal(sheet, row, item["name"], f"{item['total']} €")

        row += 1
        sheet = BillsController.createTotal(sheet, row, "TOTAL", f"{bsd['total']} €")
        row += 1
        sheet = BillsController.createTotal(sheet, row, "BANK TRANSFER", "", False)
        row += 1
        sheet = BillsController.createTotal(sheet, row, "REFUND", "", False)
        row += 1
        sheet = BillsController.createTotal(sheet, row, "PAID IN CASH", "", False)

        billName = f"bill-{event.name}-{club.name}.xlsx"
        wb.save(ResMan.web("bills", "xlsx", billName))

        return billName

    @staticmethod
    def createTotal(sheet, row, text, value, bold=True):
        sheet.merge_cells(f"B{row}:H{row}")
        sheet.cell(row, 2, text).alignment = Alignment(horizontal="center", vertical="center")
        sheet.cell(row, 9, value).alignment = Alignment(horizontal="right")
        for i in range(2, 10):
            sheet.cell(row, i).border = BillsController.medium_border
            sheet.cell(row, i).font = Font(bold=bold)
        return sheet

    @staticmethod
    def emptyRow(sheet, row):
        sheet.cell(row, 2).border = BillsController.medium_left
        sheet.cell(row, 9).border = BillsController.medium_right_thin_left
        return sheet

    @staticmethod
    def createPdf(bad, bpd, bsd, event, club):
        billName = f"bill-{event.name}-{club.name}.pdf"
        tempFile = ResMan.resources(f"temp-bill-{event.name}-{club.name}.html")

        data = BillsController.createHtml(bad, bpd, bsd, event, club)

        with open(tempFile, "w") as f:
            f.write(data)

        if (platform.system() == "Windows"):
            config = pdfkit.configuration(wkhtmltopdf = r"C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe") 
            pdfkit.from_file(tempFile, ResMan.web("bills", "pdf", billName), configuration=config)
        else:
            pdfkit.from_file(tempFile, ResMan.web("bills", "pdf", billName))

        os.remove(tempFile)
        return billName

    @staticmethod
    def createHtml(bad, bpd, bsd, event, club):
        with open(ResMan.resources("pdfTemp.html"), "r") as f:
            data = f.read()

        data = data.replace("$PORT$", str(Settings.port))
        data = data.replace("$EVENT_NAME$", event.name)
        data = data.replace("$PLACE_YEAR$", event.place.upper() + " " + event.event_start.strftime("%Y"))
        data = data.replace("$INVOICE$", "NEZAPOMENOUT")
        data = data.replace("$DATE$", datetime.now().strftime("%d.%m.%Y"))
        data = data.replace("$TO$", club.name)
        data = data.replace("$ROOMS$", BillsController.prepareRooms(bad))
        data = data.replace("$ACC_TOTAL$", str(bad["total"]) + " &#8364;")
        data = data.replace("$PACKAGES$", BillsController.preparePackages(bpd))
        data = data.replace("$PACKAGES_TOTAL$", str(bpd["total"]) + " &#8364;")
        data = data.replace("$SUMMARY$", BillsController.prepareSummary(bsd))
        data = data.replace("$SUMM_TOTAL$", str(bsd["total"]) + " &#8364;")
        return data

    @staticmethod
    def prepareRooms(bad):
        trs = ""
        for room in bad["ROOMS"]:
            trs += f"""
            <tr>
                <td colspan="2">{room["room_name"]}</td>
                <td class="center">{room["start_date"]}</td>
                <td class="center">{room["end_date"]}</td>
                <td class="center">{room["count_room"]}</td>
                <td class="center">{room["count_people"]}</td>
                <td class="center">{room["nights"]}</td>
                <td class="center">{room["price_ro"]} &#8364;</td>
                <td class="right">{room["total"]} &#8364;</td>
            </tr>
            """
        return trs

    @staticmethod
    def preparePackages(bpd):
        trs = ""
        for room in bpd["PACKAGES"]:
            trs += f"""
            <tr>
                <td colspan="2">{room["room_name"]}</td>
                <td class="center">{room["package_name"]}</td>
                <td class="center">{room["start_date"]}</td>
                <td class="center">{room["end_date"]}</td>
                <td class="center">{room["count_room"]}</td>
                <td class="center">{room["count_people"]}</td>
                <td class="center">{room["nights"]}</td>
                <td class="center">{room["price"]} &#8364;</td>
                <td class="right">{room["total"]} &#8364;</td>
            </tr>
            """
        return trs

    @staticmethod
    def prepareSummary(bsd):
        trs = ""
        for item in bsd["ITEMS"]:
            trs += f"""
            <tr>
                <td class="center"><b>{item["name"]}</b></td>
                <td class="center">{item["number"]}</td>
                <td class="center">{item["price"]} &#8364;</td>
                <td class="right">{item["total"]} &#8364;</td>
            </tr>
            """
        return trs





